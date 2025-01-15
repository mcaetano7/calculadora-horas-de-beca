import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

archivo_excel = r"C:\Users\Estudiante UCU\OneDrive - Universidad Católica del Uruguay\Escritorio\horas-de-beca.xlsx"

def config_excel():
    if not os.path.exists(archivo_excel):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Horas de beca"
        sheet.append(["Fecha", "Tarea", "Cantidad de horas", "Valor", "Total"])
        workbook.save(archivo_excel)

def guardar_excel(tarea, cantidad, valor, total):
    try:
        try:
            with open(archivo_excel, "a"):
                pass
        except PermissionError:
            messagebox.showerror("Error", "No se puede guardar con el archivo abierto.")
            return
        Workbook = load_workbook(archivo_excel)
        sheet = Workbook.active
        fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet.append([fecha_actual, tarea, cantidad, valor, total])
        Workbook.save(archivo_excel)

        os.startfile(archivo_excel)
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo guardar en el archivo de Excel: {e}")

        

def agregar_horas():
    try:
        tarea = entry_tarea.get().strip()
        horas = float(entry_horas.get())
        valor = tipo_valor.get()

        if not tarea:
            raise ValueError("Debes seleccionar una tarea")
        if horas <= 0:
            raise ValueError("Las horas deben ser mayor a 0")
        
        valor_hora = 2 if valor == "Doble" else 1
        total = horas * valor_hora

        global total_horas
        total_horas += horas
        label_total.config(text=f"Total de horas: {total_horas}")
        
        guardar_excel(tarea, horas, valor, total)
        messagebox.showinfo("Guardado", "Guardado exitosamente")
        
        entry_horas.delete(0, tk.END)
        entry_tarea.delete(0, tk.END)
        tipo_valor.set("")
    except ValueError as ve:
        messagebox.showerror("Error", str(ve))

config_excel()
total_horas = 0

ventana = tk.Tk()
ventana.title("Calculadora de horas de beca")

tk.Label(ventana, text="Tarea: ").grid(row=0, column=0, padx=10, pady=5, sticky="e")
entry_tarea = tk.Entry(ventana)
entry_tarea.grid(row=0, column=1, padx=10, pady=5)

tk.Label(ventana, text="Cantidad de horas: ").grid(row=1, column=0, padx=10, pady=5)
entry_horas = tk.Entry(ventana)
entry_horas.grid(row=1, column=1, padx=10, pady=5)

tk.Label(ventana, text="Valor:").grid(row=2, column=0, padx=10, pady=5, sticky="e")
tipo_valor = ttk.Combobox(ventana, values=["Simple", "Doble"])
tipo_valor.grid(row=2, column=1, padx=10, pady=5)
tipo_valor.set("")

btn_agregar = tk.Button(ventana, text="Agregar horas realizadas", command=agregar_horas)
btn_agregar.grid(row=3, column=0, columnspan=2, pady= 10)

label_total = tk.Label(ventana, text=f"Total de horas: {total_horas}")
label_total.grid(row=4, column=0, columnspan=2, pady=10)



ventana.mainloop()